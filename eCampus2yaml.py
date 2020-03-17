#!/usr/bin/env python
# -*- coding: utf-8 -*-

__author__ = "Raphaël LEBER"



import gitlab
import yaml
import re
from openpyxl import load_workbook
import sys



class eCampus2yaml():
    """
    Convert excel list in yaml description files
    """        


    FIRSTNAME_COL = 'A'
    FAMILYNAME_COL = 'B'
    EMAIL_COL = 'C'
    TOPIC_COL = 'D'
    GROUP_COL = 'E'

    def __init__(self):

        if len(sys.argv) > 1:
            filename = sys.argv[1]
        else:
            filename = 'students.xlsx'   

        wb = load_workbook(filename=filename) # file to open
        self.ws = wb.active

        dict_prj = {}

        for index,email in enumerate( self.ws[self.EMAIL_COL] ):

            print(index)

            if index > 0:

                firstname   =   self.ws[ self.FIRSTNAME_COL  + str(index+1)    ].value
                familyname  =   self.ws[ self.FAMILYNAME_COL + str(index+1)    ].value
                email       =   self.ws[ self.EMAIL_COL      + str(index+1)    ].value
                topic_num   =   self.ws[ self.TOPIC_COL      + str(index+1)    ].value
                team_num   =   self.ws[ self.GROUP_COL      + str(index+1)    ].value

                print ( self.FIRSTNAME_COL  + str(index+1) )
                print (firstname, "  ", familyname)


                topic = "Sujet_"+str(topic_num)
                if not topic in dict_prj:
                    dict_prj[topic] = {}

                group = "Groupe_"+str(team_num)
                if not group in dict_prj[topic] :
                    dict_prj[topic][group] = {}

                if not isinstance(dict_prj[topic][group], list) :
                    dict_prj[topic][group] = []

                dict_prj[topic][group].append([familyname, firstname, email])

        print( dict_prj )


        yaml_filename = input('\n\nEnter a file name (or press enter for default name "groups.yaml") : ')

        if len(yaml_filename) == 0 :
            yaml_filename = 'groups.yaml'         

        with open(yaml_filename, 'w') as file:
            documents = yaml.dump(dict_prj, file, allow_unicode=True)            
        


if __name__ == "__main__":
    eC2yaml = eCampus2yaml()
