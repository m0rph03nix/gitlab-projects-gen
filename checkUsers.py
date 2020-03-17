import gitlab
import yaml
import re
from openpyxl import load_workbook
import sys


class checkUsers():
    """
    Check users listed in excel file have a GitLab account
    """        

    FIRSTNAME_COL = 'A'
    FAMILYNAME_COL = 'B'
    EMAIL_COL = 'C'
    TOPIC_COL = 'D'
    GROUP_COL = 'E'    

    def __init__(self):

        if len(sys.argv) > 1:
            filename = sys.argv[1]
        else:
            filename = 'students.xlsx'

        print( filename )

        with open('token.yaml') as file:
            token = yaml.load(file, Loader=yaml.FullLoader)
            self.token = token['token_gitlab']

        self.gl = gitlab.Gitlab('https://gitlab.com', private_token=self.token)

        #print( self.gl.users.list(search='thomas.cure@cpe.fr')[0].username )

        wb = load_workbook(filename=filename) # file to open
        self.ws = wb.active

        dict_prj = {}

        color = '\x1b[0;33;41m'
        print (color+"\n")
        print ( '\x1b[0;30;41m' + "Etudiants SANS un compte : \n" + color)
        names_done = []


        for index,email in enumerate( self.ws[self.EMAIL_COL] ):

            if index > 0 :

                firstname   =   self.ws[ self.FIRSTNAME_COL  + str(index+1)    ].value
                familyname  =   self.ws[ self.FAMILYNAME_COL + str(index+1)    ].value
                email       =   self.ws[ self.EMAIL_COL      + str(index+1)    ].value

                ids = self.gl.users.list(search=email)
                if len(ids) > 0:
                    names_done.append(firstname+" "+familyname  + "  -->  " + ids[0].name)

                    #print( ids[0].name )
                    pass

                else:
                    name = firstname+" "+familyname 
                    print(color + name + '\x1b[0;33;41m')
                    if color == '\x1b[0;30;41m':
                        color = '\x1b[0;33;41m'
                    else:
                        color = '\x1b[0;30;41m'

        color = '\x1b[0m'
        print ( color + "\n\nEtudiants avec un compte (nom excel --> nom gitlab): \n" + color)
        

        for name in names_done:
            print(color + name )         


if __name__ == "__main__":
    pg = checkUsers()
